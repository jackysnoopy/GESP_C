import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

# 文件信息列表
cpp_files = [
    ("ChessboardCoverage.cpp", "棋盘覆盖问题", "使用分治法，将(2^k)*(2^k)棋盘分割为4个子棋盘，递归用L型骨牌覆盖除特殊方格外的区域"),
    ("MatchTable.cpp", "循环赛日程安排", "使用分治法生成2^k个选手的循环赛日程表，通过复制和交换矩阵元素构建完整赛程"),
    ("pack.cpp", "背包问题(0-1 Knapsack)", "动态规划求解，迭代法显示状态转移表，递归法直接计算最大价值"),
    ("Tubing.cpp", "输油管问题", "使用快速排序对y坐标排序，取中位数作为主管道最优位置以最小化总距离"),
    ("Neumann2_4_12.cpp", "冯诺依曼邻居(递归)", "递归关系 f(n)=f(n-1)+4n，求解n阶冯诺依曼邻居的元胞数"),
    ("Neumann2_3_12.cpp", "冯诺依曼邻居(通项)", "通项公式 f(n)=2n²+2n+1，直接计算n阶冯诺依曼邻居的元胞数"),
    ("FibonacciSearch.cpp", "斐波那契查找", "基于斐波那契数列的黄金分割查找算法，时间复杂度O(logn)"),
    ("HeapSort.cpp", "堆排序", "构建最大堆，反复取出堆顶元素并调整堆，时间复杂度O(nlogn)"),
    ("CountSort.cpp", "计数排序", "统计每个元素出现次数，按计数输出，非比较排序，时间复杂度O(n+k)"),
    ("BucketSort.cpp", "桶排序", "将元素分配到桶中，每个桶内插入排序后合并，适合均匀分布数据"),
    ("LinkList.cpp", "单链表", "实现单链表的创建、入队、出队、遍历等基本操作"),
    ("SqList.cpp", "顺序表", "实现顺序表的初始化、销毁、增删改查等操作"),
    ("SqStack.cpp", "顺序栈", "实现顺序栈的入栈、出栈、取栈顶等操作，支持动态扩容"),
    ("BinaryTree.cpp", "二叉树", "实现二叉树的构建、遍历、叶子节点计数、高度计算等操作"),
    ("RedBlackTree.cpp", "红黑树", "实现红黑树的插入、删除、旋转等操作，保持自平衡特性"),
    ("LinkList_with_head.cpp", "带头结点单链表", "单链表增加头结点，统一空表和非空表的操作"),
    ("HashTable.cpp", "哈希表", "除留余数法哈希函数，线性探测解决冲突，支持重建扩容"),
    ("BridgeMain.cpp", "桥接模式", "将抽象部分与实现部分分离，使二者可以独立变化"),
    ("ObserverMain.cpp", "观察者模式", "定义对象间一对多依赖，状态变化时通知所有观察者"),
    ("Singleton.cpp", "单例模式", "确保类只有一个实例，提供全局访问点"),
    ("main.cpp", "设计模式示例", "演示多种设计模式的使用：单例、抽象工厂、适配器、桥接、观察者"),
    ("FactoryMain.cpp", "抽象工厂模式", "创建一系列相关或相互依赖对象的接口，支持不同品牌工厂"),
    ("Factory.cpp", "抽象工厂实现", "工厂创建函数，根据类型返回具体工厂实例(奔驰、宝马、奥迪)"),
    ("test.cpp", "测试文件", "简单测试文件"),
    ("cell_division.cpp", "细胞分裂问题", "给定N种细胞类型和分裂率，计算使细胞总数达到m1^m2的最少时间"),
    ("001.cpp", "大整数斐波那契", "使用BigInt类计算大数斐波那契数列，避免整数溢出"),
    ("1508.cpp", "幸运数判断", "判断数是否为某个平方数的倍数，或输出下一个符合条件的数"),
    ("1503.cpp", "质因子个数判断", "判断一个数的不同质因子个数是否恰好为2"),
    ("1554.cpp", "阶乘质因数分解", "计算n!的质因数分解结果"),
    ("1556.cpp", "因子求和", "计算1到n每个数的因子个数之和"),
    ("1516.cpp", "空文件", "待实现"),
    ("1284.cpp", "质因数分解", "将整数分解为质因数乘积形式"),
    ("1157.cpp", "细胞分裂(未完成)", "细胞分裂问题代码，未完成"),
    ("P1036.cpp", "选数(NOIP 2002)", "从n个数中选k个，求和为质数的方案数(DFS回溯)"),
    ("P1075.cpp", "质因数分解(NOIP 2012)", "输出合数x的最大质因数x/p"),
    ("P1125.cpp", "幸运单词", "判断单词中字符出现次数差是否为质数")
]

# 创建工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "C++文件索引"

# 设置表头
headers = ["文件名", "问题描述", "代码思路"]
ws.append(headers)

# 设置表头样式
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for col in range(1, 4):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# 添加数据
for idx, (filename, description, approach) in enumerate(cpp_files, start=2):
    ws.cell(row=idx, column=1, value=filename)
    ws.cell(row=idx, column=2, value=description)
    ws.cell(row=idx, column=3, value=approach)

# 设置数据单元格样式
for row in range(2, len(cpp_files) + 2):
    for col in range(1, 4):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        if col == 1:
            cell.alignment = Alignment(horizontal='left', vertical='top')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# 设置列宽
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 60

# 冻结首行
ws.freeze_panes = 'A2'

# 保存文件
output_path = "/Users/jackychen/workspace/code/c++/index.xlsx"
wb.save(output_path)

print(f"Excel文件已创建: {output_path}")
print(f"共记录 {len(cpp_files)} 个C++文件")